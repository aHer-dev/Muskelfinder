"""
xlsx_to_json.py
Konvertiert die 3 XLSX-Dateien in die jeweiligen JSON-Dateien.
Führt aus mit:  python3 xlsx_to_json.py
"""

import json
import openpyxl
from pathlib import Path

DATA_DIR = Path(__file__).parent

FILES = [
    {
        "xlsx": "obere-extremitaet - + easy  + umgewandelt .xlsx",
        "json": "obere-extremitaet.json",
        "easy_sheet": "Leicht-Modus",   # enthält beide: easy (0-4) + voll (5+)
        "main_sheet": None,             # gleiche Quelle wie easy_sheet
        "region": "obere-extremitaet",
    },
    {
        "xlsx": "untere-extremitaet -+ fuß komplett umgewandelt .xlsx",
        "json": "untere-extremitaet.json",
        "easy_sheet": "Leicht-Modus",   # nur easy (0-6)
        "main_sheet": "Untere Extremität",  # nur voll (0-26)
        "region": "untere-extremitaet",
        "preserve_missing_from_existing": True,
    },
    {
        "xlsx": "wirbelsaeule-rumpf - + easy + umgewandelt.xlsx",
        "json": "wirbelsaeule.json",
        "easy_sheet": "Leicht-Modus",
        "main_sheet": None,
        "region": "wirbelsaeule",
        # Subgruppen die in kopf-hals.json ausgelagert werden
        "split_out": {
            "subgroups": {"mimikmuskulatur","kaumuskulatur","suprahyoidal","infrahyoidal","halsmuskulatur","praevertebralis"},
            "json": "kopf-hals.json",
            "region": "kopf-hals",
        },
    },
]


def parse_tags(val):
    if not val:
        return []
    return [t.strip() for t in str(val).split(",") if t.strip()]


def build_origin(row, offset):
    """Liest bis zu 4 Ursprungsteile aus der Zeile (Part/Location-Paare)."""
    parts = []
    for i in range(4):
        part_name = row[offset + i * 2]
        part_loc  = row[offset + i * 2 + 1]
        if part_loc:
            if part_name:
                parts.append({"Part": str(part_name), "Location": str(part_loc)})
            else:
                parts.append(str(part_loc))
    if not parts:
        return ""
    if len(parts) == 1:
        # Wenn nur ein Teil: gibt direkt den Wert zurück
        entry = parts[0]
        return entry["Location"] if isinstance(entry, dict) else entry
    return parts


def build_insertion(row, offset):
    """Liest bis zu 3 Ansätze."""
    vals = [row[offset + i] for i in range(3) if row[offset + i]]
    if not vals:
        return ""
    return vals[0] if len(vals) == 1 else [str(v) for v in vals]


def parse_full_row_with_easy(row):
    """
    Für obere-extremitaet und wirbelsaeule:
    Spalten 0-4: easy, Spalten 5+ : Volldata.
    Header (row 1): Name, Ursprung, Ansatz, Funktion, Innervation,
                    Name, region, subgroup, tags, difficulty, clinicalNote,
                    Joints, Movements, Segments, Image,
                    Origin_Part_1..4 (je 2 Spalten), Insertion_1..3, Function,
                    Attribution_Author, Attribution_License, Attribution_Source, Image_Source
    """
    easy = {
        "Origin":    str(row[1]) if row[1] else "",
        "Insertion": str(row[2]) if row[2] else "",
        "Function":  str(row[3]) if row[3] else "",
        "Segments":  str(row[4]) if row[4] else "",
    }

    name         = str(row[5])  if row[5]  else ""
    region       = str(row[6])  if row[6]  else ""
    subgroup     = str(row[7])  if row[7]  else ""
    tags         = parse_tags(row[8])
    difficulty   = int(row[9])  if row[9]  else None
    clinicalNote = str(row[10]) if row[10] else ""
    joints       = str(row[12]) if row[12] else ""
    movements    = str(row[13]) if row[13] else ""
    segments     = str(row[14]) if row[14] else ""
    image        = str(row[15]) if row[15] else ""
    # Origin: cols 16-23 (4 Parts × 2)
    origin       = build_origin(row, 16)
    # Insertion: cols 24-26
    insertion    = build_insertion(row, 24)
    function_    = str(row[27]) if row[27] else ""
    attr_author  = str(row[28]) if row[28] else ""
    attr_license = str(row[29]) if row[29] else ""
    attr_source  = str(row[30]) if row[30] else ""
    image_source = str(row[31]) if row[31] else ""

    entry = {
        "Name": name,
        "Joints": joints,
        "Movements": movements,
        "Segments": segments,
        "Origin": origin,
        "Insertion": insertion,
        "Function": function_,
        "Image": image,
        "region": region,
        "subgroup": subgroup,
        "tags": tags,
        "difficulty": difficulty,
        "clinicalNote": clinicalNote,
        "functionalChain": "",
        "Attribution": {
            "Author": attr_author,
            "AuthorUrl": "",
            "License": attr_license,
            "LicenseUrl": attr_source,
            "SourceUrl": image_source,
        },
        "easy": easy,
    }
    return entry


def parse_full_row_no_easy(row):
    """
    Für untere-extremitaet (Hauptblatt): keine easy-Spalten.
    Cols: Name, region, subgroup, tags, difficulty, clinicalNote,
          Joints, Movements, Segments, Image,
          Origin_Part_1..4 (je 2 Spalten), Insertion_1..3, Function,
          Attribution_Author, Attribution_License, Attribution_Source, Image_Source
    """
    name         = str(row[0])  if row[0]  else ""
    region       = str(row[1])  if row[1]  else ""
    subgroup     = str(row[2])  if row[2]  else ""
    tags         = parse_tags(row[3])
    difficulty   = int(row[4])  if row[4]  else None
    clinicalNote = str(row[5])  if row[5]  else ""
    joints       = str(row[7])  if row[7]  else ""
    movements    = str(row[8])  if row[8]  else ""
    segments     = str(row[9])  if row[9]  else ""
    image        = str(row[10]) if row[10] else ""
    origin       = build_origin(row, 11)
    insertion    = build_insertion(row, 19)
    function_    = str(row[22]) if row[22] else ""
    attr_author  = str(row[23]) if row[23] else ""
    attr_license = str(row[24]) if row[24] else ""
    attr_source  = str(row[25]) if row[25] else ""
    image_source = str(row[26]) if row[26] else ""

    return {
        "Name": name,
        "Joints": joints,
        "Movements": movements,
        "Segments": segments,
        "Origin": origin,
        "Insertion": insertion,
        "Function": function_,
        "Image": image,
        "region": region,
        "subgroup": subgroup,
        "tags": tags,
        "difficulty": difficulty,
        "clinicalNote": clinicalNote,
        "functionalChain": "",
        "Attribution": {
            "Author": attr_author,
            "AuthorUrl": "",
            "License": attr_license,
            "LicenseUrl": attr_source,
            "SourceUrl": image_source,
        },
    }


def parse_easy_only_row(row):
    """
    Für untere-extremitaet Leicht-Modus-Sheet:
    Cols: Name, Ursprung, Ansatz, Funktion, Innervation, clinicalNote
    """
    return {
        "name": str(row[0]) if row[0] else "",
        "easy": {
            "Origin":    str(row[1]) if row[1] else "",
            "Insertion": str(row[2]) if row[2] else "",
            "Function":  str(row[3]) if row[3] else "",
            "Segments":  str(row[4]) if row[4] else "",
        }
    }


def is_section_header(name):
    """Erkennt Trennzeilen wie '── HÜFTE ──'."""
    return name.startswith("──") or name.startswith("─")


def load_existing_entries(json_name):
    """Lädt vorhandene JSON-Einträge, um App-spezifische Felder zu bewahren."""
    path = DATA_DIR / json_name
    if not path.exists():
        return {}

    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError):
        return {}

    return {
        str(entry.get("Name")): entry
        for entry in data.get("Sheet1", [])
        if entry.get("Name")
    }


def merge_existing_entry(entry, existing):
    """Behält Bildpfade, Galerien und leere Metadaten aus bestehenden JSONs bei."""
    if not existing:
        return entry

    if not entry.get("Image") and existing.get("Image"):
        entry["Image"] = existing["Image"]

    if not entry.get("Images"):
        if existing.get("Images"):
            entry["Images"] = existing["Images"]
        elif entry.get("Image"):
            entry["Images"] = [entry["Image"]]
    elif entry.get("Image") and entry["Image"] not in entry["Images"]:
        entry["Images"] = [entry["Image"], *entry["Images"]]

    old_attr = existing.get("Attribution") or {}
    new_attr = entry.get("Attribution") or {}
    for key in ("AuthorUrl", "SourceUrl", "LicenseUrl"):
        if not new_attr.get(key) and old_attr.get(key):
            new_attr[key] = old_attr[key]
    entry["Attribution"] = new_attr

    return entry


def prepare_output_entries(json_name, muscles, preserve_missing=False):
    """Mergt bestehende App-Felder in neu erzeugte Einträge und bewahrt optional fehlende Alt-Einträge."""
    existing_map = load_existing_entries(json_name)
    prepared = []
    seen_names = set()

    for muscle in muscles:
        name = str(muscle.get("Name") or "").strip()
        if not name:
            continue
        prepared.append(merge_existing_entry(muscle, existing_map.get(name)))
        seen_names.add(name)

    if preserve_missing:
        for name, existing in existing_map.items():
            if name not in seen_names:
                prepared.append(existing)

    return prepared


def convert(cfg):
    xlsx_path = DATA_DIR / cfg["xlsx"]
    wb = openpyxl.load_workbook(xlsx_path)

    muscles = []

    if cfg["main_sheet"] is None:
        # obere + wirbelsaeule: Leicht-Modus enthält alles
        ws = wb[cfg["easy_sheet"]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 3:
                continue  # Header-Zeilen überspringen
            name_easy = str(row[0]) if row[0] else ""
            name_full = str(row[5]) if row[5] else ""
            name = name_full or name_easy
            if not name or is_section_header(name):
                continue
            entry = parse_full_row_with_easy(row)
            muscles.append(entry)
    else:
        # untere: Hauptblatt für Volldata, Leicht-Modus für easy
        ws_main = wb[cfg["main_sheet"]]
        ws_easy = wb[cfg["easy_sheet"]]

        # Volldata einlesen
        full_data = {}
        for i, row in enumerate(ws_main.iter_rows(values_only=True)):
            if i < 3:
                continue
            name = str(row[0]) if row[0] else ""
            if not name or is_section_header(name):
                continue
            full_data[name] = parse_full_row_no_easy(row)

        # Easy-Daten einlesen und mergen
        easy_data = {}
        for i, row in enumerate(ws_easy.iter_rows(values_only=True)):
            if i < 3:
                continue
            name = str(row[0]) if row[0] else ""
            if not name or is_section_header(name):
                continue
            parsed = parse_easy_only_row(row)
            easy_data[parsed["name"]] = parsed["easy"]

        for name, entry in full_data.items():
            if name in easy_data:
                entry["easy"] = easy_data[name]
            muscles.append(entry)

    # Ggf. Subgruppen in separate Datei auslagern
    if "split_out" in cfg:
        split = cfg["split_out"]
        split_muscles = [m for m in muscles if m.get("subgroup","") in split["subgroups"]]
        main_muscles  = [m for m in muscles if m.get("subgroup","") not in split["subgroups"]]
        for m in split_muscles:
            m["region"] = split["region"]
        split_muscles = prepare_output_entries(split["json"], split_muscles)
        with open(DATA_DIR / split["json"], "w", encoding="utf-8") as f:
            json.dump({"Sheet1": split_muscles}, f, ensure_ascii=False, indent=2)
        print(f"  ✓ {split['json']}  ({len(split_muscles)} Muskeln)")
        muscles = main_muscles

    # In JSON schreiben
    muscles = prepare_output_entries(
        cfg["json"],
        muscles,
        preserve_missing=cfg.get("preserve_missing_from_existing", False),
    )
    out_path = DATA_DIR / cfg["json"]
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"Sheet1": muscles}, f, ensure_ascii=False, indent=2)

    print(f"  ✓ {cfg['json']}  ({len(muscles)} Muskeln)")


if __name__ == "__main__":
    print("Konvertiere XLSX → JSON ...")
    for cfg in FILES:
        convert(cfg)
    print("Fertig!")
